"""Module providing openpyxl extensions for framework"""

from openpyxl import Workbook, load_workbook


class ExcelFramework:
    """AI is creating summary for"""

    def __init__(self, filename=None):
        """__init__ declaration for objects

        Args:
            filename (_ZipFileFileProtocol, optional): filename of the file. Defaults to None.
        """
        if filename:
            self.workbook = load_workbook(filename)
        else:
            self.workbook = Workbook()
            self.workbook.create_sheet("Sheet1")

    def add_data(self, sheet_name: str, data: list, start_row=1, start_column=1):
        """Adds data to excel.

        Args:
            sheet_name (string): name of the sheet name for the data to be added on
            data (list): the following data to be added in the excel file, in list format.
            start_row (int, optional): row number to start putting data on. Defaults to 1.
            start_column (int, optional): column number to start putting data on. Defaults to 1.
        """
        sheet = self.workbook[sheet_name]
        for row_index, row_data in enumerate(data, start=start_row):
            for col_index, cell_data in enumerate(row_data, start=start_column):
                sheet.cell(row=row_index, column=col_index, value=cell_data)

    def save_workbook(self, filename: str):
        """save_workbook AI is creating summary for save_workbook

        Args:
            filename (_ZipFileFileProtocol): filename of the file.
        """
        self.workbook.save(filename)

    def load_workbook(self, filename: str):
        """
        load_workbook: method to use when loading an Excel Filename

        Args:
            filename (_ZipFileFileProtocol): filename of the file.
        """
        self.workbook = load_workbook(filename)

    def __cell_is_occupied(self, sheet_name: str, cell_name: str):
        sheet = self.workbook[sheet_name]
        cell = sheet[cell_name]
        return cell.value is not None

    def add_data_to_empty_row_cell(self, sheet_name: str, data: list):
        """_summary_: This method identifies the nearest available row to add
            data in an Excel workbook when the first few rows are already occupied.

        Args:
            sheet_name (str): _description_
            data (list): _description_
        """
        row_num = 1

        while self.__cell_is_occupied(sheet_name, f"A{row_num}"):
            row_num = row_num + 1

        self.add_data(sheet_name, data, row_num)
