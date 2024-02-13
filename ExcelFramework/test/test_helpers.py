"""Modules needed for test_helpers"""

import os
from openpyxl import load_workbook
from ..package.excel_framework import ExcelFramework
from .static_constants import (
    SHEET_NAMES_DIFFERENT,
    DIMENSIONS_SHEET_NAME_DIFFERENT,
    CELL_IN_SHEET_NAME_DIFFERENT,
    EXCEL_FILES_EQUAL,
)


class TestHelpers:
    """_summary_: methods that help execute test_excel_framework"""

    @staticmethod
    def validate_excel_files(file1, file2):
        """_summary_: checks if two excel files are equal

        Args:
            file1 (string): excel filename to be tested
            file2 (string): expected excel filename

        Returns:
            _type_: boolean, string
        """
        workbook1 = load_workbook(file1)
        workbook2 = load_workbook(file2)

        # Get all sheet names from both workbooks
        sheet_names1, sheet_names2 = workbook1.sheetnames, workbook2.sheetnames

        # Check if sheet names are the same
        if sheet_names1 != sheet_names2:
            return False, SHEET_NAMES_DIFFERENT

        # Iterate over each sheet and compare cell values
        for sheet_name in sheet_names1:
            sheet1, sheet2 = workbook1[sheet_name], workbook2[sheet_name]

            # Check if dimensions are the same
            if (
                sheet1.max_row != sheet2.max_row
                or sheet1.max_column != sheet2.max_column
            ):
                return False, DIMENSIONS_SHEET_NAME_DIFFERENT.replace(
                    r"{sheet_name}", sheet_name
                )

            # Compare cell values
            for row in range(1, sheet1.max_row + 1):
                for col in range(1, sheet1.max_column + 1):
                    if (
                        sheet1.cell(row=row, column=col).value
                        != sheet2.cell(row=row, column=col).value
                    ):
                        return (
                            False,
                            CELL_IN_SHEET_NAME_DIFFERENT,
                        )

        # If all checks pass, the Excel files are equal
        return True, EXCEL_FILES_EQUAL

    @staticmethod
    def create_workbook(sheet_name, data, filename_to_save, filename_to_open=None):
        """_summary_: reusable method to create and save workbook

        Args:
            sheet_name (string): sheet name where to insert the data
            data (list): a list of a list of data containing the
            filename_to_save (string): name of the file for it to be saved.
            filename_to_open (string, optional): name of the file to be opened if parameter
                is not empty. Defaults to None.
        """
        # excel_framework = ExcelFramework(filename_to_open)
        excel_framework = ExcelFramework(filename_to_open)
        excel_framework.add_data(sheet_name, data)
        excel_framework.save_workbook(filename_to_save)

    @staticmethod
    def delete_file(file_path):
        """_summary_: deletes file if it exists

        Args:
            file_path (string): directory to file to be deleted
        """
        if os.path.exists(file_path):
            os.remove(file_path)
